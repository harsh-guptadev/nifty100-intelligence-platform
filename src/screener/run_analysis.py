import os
import sqlite3
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.screener.engine import (
    get_db_connection,
    get_screener_data,
    compute_sector_relative_scores,
    update_composite_scores_in_db,
    run_preset_screener,
    load_screener_config,
    map_cagr_flag_to_numeric
)
from src.analytics.peer import (
    populate_peer_percentiles,
    get_company_peer_group
)

# Output Paths
SCREENER_OUT_PATH = "output/screener_output.xlsx"
PEER_OUT_PATH = "output/peer_comparison.xlsx"
RADAR_CHARTS_DIR = "reports/radar_charts"

# Styling Colors
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="Calibri", size=11, color="006100", bold=True)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="Calibri", size=11, color="9C0006", bold=True)

YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT = Font(name="Calibri", size=11, color="9C6500", bold=True)

GOLD_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
GOLD_FONT = Font(name="Calibri", size=11, color="000000", bold=True)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, color="FFFFFF", bold=True)

MEDIAN_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
MEDIAN_FONT = Font(name="Calibri", size=11, color="000000", bold=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# DESIGN DEVIATION FROM SPEC:
# The Day 19 specification requested a "single-metric standalone chart" for companies with no peer group.
# However, the implementation intentionally reuses the standard 8-axis polar/radar chart layout for peerless companies,
# substituting the Nifty 100 universe average (50th percentile rank) as the reference benchmark instead.
# This design decision was made to ensure visual consistency and standardized reporting formatting across all 92 company reports.
def create_radar_charts(df_all: pd.DataFrame, df_pct: pd.DataFrame):
    """Generates and exports PNG radar charts for all companies."""
    os.makedirs(RADAR_CHARTS_DIR, exist_ok=True)
    
    # Group peer percentiles by company and year (latest)
    # The metrics to rank are: ROE, ROCE, Net Profit Margin, D/E, FCF, PAT CAGR 5yr, Revenue CAGR 5yr, Composite Score
    # We will fetch latest percentile ranks for each company
    conn = get_db_connection()
    try:
        df_db_pct = pd.read_sql_query("SELECT * FROM peer_percentiles", conn)
        df_peer_groups = pd.read_sql_query("SELECT * FROM peer_groups", conn)
    finally:
        conn.close()
        
    # Get latest year for each company (latest with non-null return_on_equity_pct, falling back to absolute latest)
    df_valid = df_all[df_all['return_on_equity_pct'].notna()]
    if not df_valid.empty:
        latest_indices = df_valid.groupby("company_id")["year"].idxmax()
        missing_cos = set(df_all['company_id']) - set(df_valid['company_id'])
        if missing_cos:
            df_missing = df_all[df_all['company_id'].isin(missing_cos)]
            missing_indices = df_missing.groupby("company_id")["year"].idxmax()
            all_indices = pd.concat([latest_indices, missing_indices])
        else:
            all_indices = latest_indices
        df_latest_yr = df_all.loc[all_indices].copy()
    else:
        df_latest_yr = df_all[df_all.groupby("company_id")["year"].transform("max") == df_all["year"]].copy()
    
    # We need: ROE, ROCE, NPM, D/E, FCF score, PAT CAGR 5yr, Revenue CAGR 5yr, Composite Score
    # Let's map metric names to database peer_percentiles names
    metrics_list = ["ROE", "ROCE", "Net Profit Margin", "D/E", "FCF", "PAT CAGR 5yr", "Revenue CAGR 5yr", "Composite Score"]
    
    # Pre-build averages per peer group
    peer_group_averages = {}
    universe_averages = {m: 50.0 for m in metrics_list} # Default universe average percentile rank is 50%
    universe_averages["Composite Score"] = df_latest_yr["composite_quality_score"].mean()

    # Calculate average percentile rank and composite score for each peer group
    for pg_name in df_peer_groups["peer_group_name"].unique():
        pg_cos = df_peer_groups[df_peer_groups["peer_group_name"] == pg_name]["company_id"].unique()
        df_pg_latest = df_latest_yr[df_latest_yr["company_id"].isin(pg_cos)]
        
        avgs = {}
        for m in metrics_list:
            if m == "Composite Score":
                avgs[m] = df_pg_latest["composite_quality_score"].mean()
            else:
                # Average percentile rank * 100 for display
                sub_pct = df_db_pct[(df_db_pct["peer_group_name"] == pg_name) & (df_db_pct["metric"] == m)]
                # Filter to latest year for each company in the sub_pct
                latest_pcts = []
                for c in pg_cos:
                    c_pct = sub_pct[sub_pct["company_id"] == c]
                    if not c_pct.empty:
                        # Get latest
                        latest_pcts.append(c_pct.sort_values("year").iloc[-1]["percentile_rank"] * 100.0)
                avgs[m] = np.mean(latest_pcts) if latest_pcts else 50.0
        peer_group_averages[pg_name] = avgs

    # Also compute universe averages from all companies for peerless companies
    for m in metrics_list:
        if m != "Composite Score":
            # Average percentile rank of all companies is 50%
            universe_averages[m] = 50.0

    print("Generating radar charts...")
    for _, row in df_latest_yr.iterrows():
        comp_id = row['company_id']
        year = row['year']
        pg_name = get_company_peer_group(comp_id)
        
        # Get company's scores
        comp_scores = []
        for m in metrics_list:
            if m == "Composite Score":
                comp_scores.append(row['composite_quality_score'])
            else:
                # Get from peer_percentiles
                c_pct = df_db_pct[(df_db_pct["company_id"] == comp_id) & (df_db_pct["metric"] == m) & (df_db_pct["year"] == year)]
                if not c_pct.empty:
                    comp_scores.append(c_pct.iloc[0]["percentile_rank"] * 100.0)
                else:
                    comp_scores.append(0.0) # Fallback

        # Determine reference averages (peer group average or universe average)
        if pg_name != "No peer group assigned":
            ref_averages = [peer_group_averages[pg_name][m] for m in metrics_list]
            ref_label = "Peer Group Avg"
            title_suffix = f"({pg_name} Peer Group)"
        else:
            # DESIGN DEVIATION: Reusing 8-axis format with Nifty 100 average for layout consistency instead of a separate single-metric chart type.
            ref_averages = [universe_averages[m] for m in metrics_list]
            ref_label = "Nifty 100 Avg"
            title_suffix = "(No Peer Group)"
            
        # Draw Radar/Polar Plot
        labels = ['ROE', 'ROCE', 'NPM', 'D/E', 'FCF', 'PAT CAGR', 'Rev CAGR', 'Composite']
        num_vars = len(labels)
        angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
        
        # Complete the loop
        angles += angles[:1]
        comp_scores += comp_scores[:1]
        ref_averages += ref_averages[:1]
        
        fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
        
        # Draw company polygon
        ax.fill(angles, comp_scores, color='#1F4E79', alpha=0.25, label=comp_id)
        ax.plot(angles, comp_scores, color='#1F4E79', linewidth=2)
        
        # Draw benchmark line
        ax.plot(angles, ref_averages, color='#D66011', linewidth=1.5, linestyle='--', label=ref_label)
        
        # Fixed 0-100 scale
        ax.set_ylim(0, 100)
        
        # Clean background and label styling
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(labels, fontsize=10, weight='bold')
        
        ax.tick_params(colors='#555555')
        ax.grid(color='#E5E5E5')
        
        plt.legend(loc='upper right', bbox_to_anchor=(1.2, 1.15))
        plt.title(f"{comp_id} Financial Health Radar\n{title_suffix}", size=12, color='#1F4E79', y=1.1, weight='bold')
        
        save_path = os.path.join(RADAR_CHARTS_DIR, f"{comp_id}_radar.png")
        plt.tight_layout()
        plt.savefig(save_path, dpi=120)
        plt.close()

def generate_screener_xlsx(df_all: pd.DataFrame):
    """Generates output/screener_output.xlsx with 6 sheets, color-coded cells."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    config = load_screener_config()
    presets = config.get("presets", {})
    
    # 20 KPI columns mapping (for display and export)
    export_columns_map = {
        "company_id": "Ticker",
        "company_name": "Company Name",
        "broad_sector": "Broad Sector",
        "year": "Year",
        "market_cap_crore": "Market Cap (Cr)",
        "sales": "Sales (Cr)",
        "net_profit": "Net Profit (Cr)",
        "return_on_equity_pct": "ROE (%)",
        "roce_percentage": "ROCE (%)",
        "net_profit_margin_pct": "Net Profit Margin (%)",
        "operating_profit_margin_pct": "OPM (%)",
        "debt_to_equity": "Debt to Equity",
        "interest_coverage": "Interest Coverage",
        "asset_turnover": "Asset Turnover",
        "free_cash_flow_cr": "FCF (Cr)",
        "cash_from_operations_cr": "CFO (Cr)",
        "earnings_per_share": "EPS",
        "pe_ratio": "P/E Ratio",
        "pb_ratio": "P/B Ratio",
        "dividend_yield_pct": "Dividend Yield (%)",
        "revenue_cagr_5yr": "Revenue CAGR 5yr",
        "pat_cagr_5yr": "PAT CAGR 5yr",
        "composite_quality_score": "Composite Score"
    }
    
    # Presets mapping to their filter criteria columns for highlight logic
    preset_highlight_cols = {
        "Quality Compounder": {
            "return_on_equity_pct": "return_on_equity_pct",
            "debt_to_equity": "debt_to_equity",
            "free_cash_flow_cr": "free_cash_flow_cr",
            "revenue_cagr_5yr": "revenue_cagr_5yr"
        },
        "Value Pick": {
            "pe_ratio": "pe_ratio",
            "pb_ratio": "pb_ratio",
            "debt_to_equity": "debt_to_equity",
            "dividend_yield_pct": "dividend_yield_pct"
        },
        "Growth Accelerator": {
            "pat_cagr_5yr": "pat_cagr_5yr",
            "revenue_cagr_5yr": "revenue_cagr_5yr",
            "debt_to_equity": "debt_to_equity"
        },
        "Dividend Champion": {
            "dividend_yield_pct": "dividend_yield_pct",
            "dividend_payout_ratio_pct": "dividend_payout_ratio_pct",
            "free_cash_flow_cr": "free_cash_flow_cr"
        },
        "Debt-Free Blue Chip": {
            "debt_to_equity": "debt_to_equity",
            "return_on_equity_pct": "return_on_equity_pct",
            "sales": "sales"
        },
        "Turnaround Watch": {
            "revenue_cagr_3yr": "revenue_cagr_3yr",
            "free_cash_flow_cr": "free_cash_flow_cr",
            "de_declining_yoy": "de_declining_yoy"
        }
    }

    for preset_name in presets.keys():
        print(f"Running screener preset: {preset_name}...")
        df_preset = run_preset_screener(preset_name, df_all)
        
        ws = wb.create_sheet(title=preset_name)
        ws.views.sheetView[0].showGridLines = True
        
        # Prepare subset for Excel
        export_df = df_preset[[col for col in export_columns_map.keys() if col in df_preset.columns]].copy()
        # Rename columns for presentation
        export_df = export_df.rename(columns=export_columns_map)
        
        # Write headers
        headers = list(export_df.columns)
        ws.append(headers)
        
        # Format headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        
        # Write rows
        highlight_mapping = preset_highlight_cols.get(preset_name, {})
        
        # For mapping df indices back to Openpyxl rows
        for r_idx, (orig_idx, row) in enumerate(df_preset.iterrows(), start=2):
            row_data = [row.get(col) for col in export_columns_map.keys() if col in df_preset.columns]
            ws.append(row_data)
            
            # Format row cells
            filter_matches = row.get('_filter_matches', {})
            
            for c_idx, col_name in enumerate(export_columns_map.keys(), start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = THIN_BORDER
                
                # Check alignment
                if col_name in ["company_id", "company_name", "broad_sector", "year"]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                
                # Format numbers
                val = cell.value
                if val is not None and isinstance(val, (int, float)):
                    if col_name in ["market_cap_crore", "sales", "net_profit", "free_cash_flow_cr", "cash_from_operations_cr"]:
                        cell.number_format = '#,##0.0'
                    elif col_name in ["return_on_equity_pct", "roce_percentage", "net_profit_margin_pct", "operating_profit_margin_pct", "pe_ratio", "pb_ratio", "dividend_yield_pct", "revenue_cagr_5yr", "pat_cagr_5yr", "composite_quality_score"]:
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '0.00'
                
                # Apply green/red conditional fills for filter matching columns
                if col_name in highlight_mapping:
                    filter_col = highlight_mapping[col_name]
                    is_match = filter_matches.get(filter_col)
                    if is_match == True:
                        cell.fill = GREEN_FILL
                        cell.font = GREEN_FONT
                    elif is_match == False:
                        cell.fill = RED_FILL
                        cell.font = RED_FONT
                        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save Workbook
    wb.save(SCREENER_OUT_PATH)
    print(f"Screener output saved to: {SCREENER_OUT_PATH}")

def generate_peer_comparison_xlsx(df_all: pd.DataFrame):
    """Generates output/peer_comparison.xlsx with 11 sheets, percentile color coding, benchmark highlighted."""
    conn = get_db_connection()
    try:
        # Load peer percentiles & peer groups
        df_db_pct = pd.read_sql_query("SELECT * FROM peer_percentiles", conn)
        df_peer_groups = pd.read_sql_query("SELECT * FROM peer_groups", conn)
    finally:
        conn.close()

    # Get latest year for each company (latest with non-null return_on_equity_pct, falling back to absolute latest)
    df_valid = df_all[df_all['return_on_equity_pct'].notna()]
    if not df_valid.empty:
        latest_indices = df_valid.groupby("company_id")["year"].idxmax()
        missing_cos = set(df_all['company_id']) - set(df_valid['company_id'])
        if missing_cos:
            df_missing = df_all[df_all['company_id'].isin(missing_cos)]
            missing_indices = df_missing.groupby("company_id")["year"].idxmax()
            all_indices = pd.concat([latest_indices, missing_indices])
        else:
            all_indices = latest_indices
        df_latest_yr = df_all.loc[all_indices].copy()
    else:
        df_latest_yr = df_all[df_all.groupby("company_id")["year"].transform("max") == df_all["year"]].copy()

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    # Columns to export
    # Standard columns: Ticker, Company Name, and 10 metrics + 10 percentile ranks
    metrics_list = [
        ("ROE (%)", "return_on_equity_pct", "ROE"),
        ("ROCE (%)", "roce_percentage", "ROCE"),
        ("NPM (%)", "net_profit_margin_pct", "Net Profit Margin"),
        ("D/E", "debt_to_equity", "D/E"),
        ("FCF (Cr)", "free_cash_flow_cr", "FCF"),
        ("PAT CAGR 5Y", "pat_cagr_numeric", "PAT CAGR 5yr"),
        ("Rev CAGR 5Y", "rev_cagr_numeric", "Revenue CAGR 5yr"),
        ("EPS CAGR 5Y", "eps_cagr_numeric", "EPS CAGR 5yr"),
        ("ICR", "interest_coverage", "Interest Coverage"),
        ("Asset Turnover", "asset_turnover", "Asset Turnover")
    ]
    
    # Pre-calculate numeric equivalents for CAGR columns to calculate medians correctly
    df_latest_yr['rev_cagr_numeric'] = df_latest_yr.apply(lambda r: map_cagr_flag_to_numeric(r['revenue_cagr_5yr'], r['revenue_cagr_5yr_flag']), axis=1)
    df_latest_yr['pat_cagr_numeric'] = df_latest_yr.apply(lambda r: map_cagr_flag_to_numeric(r['pat_cagr_5yr'], r['pat_cagr_5yr_flag']), axis=1)
    df_latest_yr['eps_cagr_numeric'] = df_latest_yr.apply(lambda r: map_cagr_flag_to_numeric(r['eps_cagr_5yr'], r['eps_cagr_5yr_flag']), axis=1)

    for peer_group in sorted(df_peer_groups["peer_group_name"].unique()):
        print(f"Generating peer sheet: {peer_group}...")
        ws = wb.create_sheet(title=peer_group)
        ws.views.sheetView[0].showGridLines = True
        
        # Get companies in peer group
        pg_cos = df_peer_groups[df_peer_groups["peer_group_name"] == peer_group]
        df_pg = df_latest_yr[df_latest_yr["company_id"].isin(pg_cos["company_id"])].copy()
        
        # Merge benchmark flag
        df_pg = df_pg.merge(pg_cos[["company_id", "is_benchmark"]], on="company_id", how="left")
        
        # Set up headers: Ticker, Company, then for each metric: Metric, Percentile Rank
        headers = ["Ticker", "Company Name"]
        for label, _, _ in metrics_list:
            headers.extend([label, f"{label} Rank"])
            
        ws.append(headers)
        
        # Format Header
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            
        # Write rows
        data_rows_start = 2
        for r_idx, (_, row) in enumerate(df_pg.iterrows(), start=data_rows_start):
            comp_id = row['company_id']
            yr = row['year']
            is_bench = row.get('is_benchmark', 0) == 1
            
            row_data = [comp_id, row['company_name']]
            
            # For each metric, write the value and its percentile rank from database
            for _, col_name, metric_db_name in metrics_list:
                val = row.get(col_name)
                # Fetch rank
                pct_row = df_db_pct[(df_db_pct["company_id"] == comp_id) & (df_db_pct["metric"] == metric_db_name) & (df_db_pct["year"] == yr)]
                rank_val = pct_row.iloc[0]["percentile_rank"] if not pct_row.empty else np.nan
                
                row_data.extend([val, rank_val])
                
            ws.append(row_data)
            
            # Format row cells
            row_fill = GOLD_FILL if is_bench else None
            row_font = GOLD_FONT if is_bench else None
            
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = THIN_BORDER
                
                # Basic alignments
                if c_idx <= 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    
                # Format numbers
                val_cell = cell.value
                if val_cell is not None and isinstance(val_cell, (int, float)):
                    # Odd columns (c_idx > 2) are values, even columns are ranks
                    if c_idx % 2 == 1:
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '0.0%' # format as percentage rank
                        
                # If row is benchmark, apply gold highlight
                if is_bench:
                    cell.fill = row_fill
                    cell.font = row_font
                    
                # Apply green/yellow/red color coding on percentile rank cells (even columns > 2)
                # Skip overriding benchmark font/fill if it is the benchmark company
                if c_idx > 2 and c_idx % 2 == 0 and not is_bench:
                    rank_pct = cell.value
                    if rank_pct is not None and isinstance(rank_pct, (int, float)):
                        if rank_pct >= 0.75:
                            cell.fill = GREEN_FILL
                            cell.font = GREEN_FONT
                        elif rank_pct <= 0.25:
                            cell.fill = RED_FILL
                            cell.font = RED_FONT
                        else:
                            cell.fill = YELLOW_FILL
                            cell.font = YELLOW_FONT

        # Add Median row at the bottom
        median_row_idx = len(df_pg) + data_rows_start
        median_data = ["Median", "Peer Group Median"]
        
        for label, col_name, _ in metrics_list:
            # Median of the metric column
            col_vals = df_pg[col_name].dropna()
            # Exclude non-numeric values
            col_vals = col_vals[col_vals.apply(lambda x: isinstance(x, (int, float)))]
            med = col_vals.median() if not col_vals.empty else np.nan
            
            median_data.extend([med, np.nan]) # Write nan for percentile rank cell
            
        ws.append(median_data)
        
        # Style the Median row
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=median_row_idx, column=c_idx)
            cell.fill = MEDIAN_FILL
            cell.font = MEDIAN_FONT
            cell.border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='double', color='000000'), # double top border for medians
                bottom=Side(style='double', color='000000')
            )
            
            if c_idx <= 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                
            val_cell = cell.value
            if val_cell is not None and isinstance(val_cell, (int, float)):
                cell.number_format = '0.00'
                
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(PEER_OUT_PATH)
    print(f"Peer comparison output saved to: {PEER_OUT_PATH}")

def main():
    print("=== STARTING SPRINT 3 RUN ANALYSIS ===")
    
    # 1. Update composite quality scores in database
    print("Step 1: Calculating and updating composite scores in database...")
    updated_scores = update_composite_scores_in_db()
    print(f"Updated {updated_scores} rows in financial_ratios.")

    # 2. Populate peer percentile rankings in database
    print("Step 2: Calculating and populating peer percentile rankings in database...")
    inserted_pcts = populate_peer_percentiles()
    print(f"Inserted {inserted_pcts} rows in peer_percentiles.")
    
    # 3. Pull scored data for report export
    df_scored = get_screener_data()
    df_scored = compute_sector_relative_scores(df_scored)
    
    # 4. Generate Screener output report
    print("Step 3: Generating output/screener_output.xlsx...")
    generate_screener_xlsx(df_scored)
    
    # 5. Generate Peer Comparison output report
    print("Step 4: Generating output/peer_comparison.xlsx...")
    generate_peer_comparison_xlsx(df_scored)
    
    # 6. Generate Radar charts
    print("Step 5: Generating reports/radar_charts/*.png...")
    create_radar_charts(df_scored, df_scored)
    
    print("=== SPRINT 3 RUN ANALYSIS COMPLETED SUCCESSFULLY ===")

if __name__ == "__main__":
    main()
